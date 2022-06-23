"""
作品名稱     ： tkinter_SQL_ERP
__author__ ： Ning 甯詠城
"""

import pymysql as MySQLdb                #  pip install MySQLdb    # MySQL
# -*- coding: utf-8 -*-
#!/usr/bin/python
import tkinter as tk
from tkinter import ttk
from tkinter import Tk, Frame, Menu, Button
from tkinter import LEFT, TOP, X, FLAT, RAISED
from tkinter.scrolledtext import ScrolledText
from tkinter import *
from PIL import Image, ImageTk
import xlwt
import csv
import pandas as pd
import xlsxwriter
import openpyxl
# encoding: utf-8



# 連接資料庫
db = MySQLdb.connect(host="127.0.0.1",   #  連接到本身的電腦IP
                     user="admin",       #  MySQL/PHPMyAdmin 新增的 用戶
                     passwd="admin",
                     db="mydatabase")    #  MySQL/PHPMyAdmin 新增的 資料庫
cursor = db.cursor()


def SQLDataPrint(list1):
    b = []
    x = 0
    for row in list1:
        a = []
        y = 0
        str1 = ""
        for col in row:
            a.append(str(col))
            str1 = str1 + str(col) + " , "
            y = y + 1
        print(str1)
        x = x + 1
        b.append(a)
    print(b)
    return b


###新增excel
caseData=["Id","LastName","FirstName","Address","City","Age"]             #讓首行永遠有caseDate內的資料
def write_excel(caseData,b):
    write = xlsxwriter.Workbook('write.xlsx') # 新增一個新的excel
    write2 = write.add_worksheet("My sheet")

    for x in range(0,4):#上下欄位
        if x==0:
            for i in range(len(caseData)):
                write2.write(x,i,caseData[i])
    write.close()
    book = openpyxl.load_workbook('write.xlsx', data_only=True)
    sheet = book.active

    for x in range(len(b)):                                         #for矩陣跑輸入的list1
        row = sheet.max_row + 1                                     # 讀取最大的row，這樣新的資料都能放入新的欄位中
        for y in range(len(b[0])):
            sheet.cell(row=row, column=y + 1).value = b[x][y]       #將輸入的值放入excel欄位中
    book.save('write.xlsx')                                         #save這次加入的東西
    write.close()



####### 0. 視窗初始化
win = tk.Tk()
win.wm_title("ERP訂單管理系統")                               # ERP enterprise resource planning
win.minsize(width=666, height=480)                          # 最小的視窗
win.maxsize(width=666, height=480)                          # 最大的視窗
win.resizable(width=False, height=False)                    # 是否可以調整


def view():                                                         # 顯示SQL

    sql = "SELECT * FROM `persons`"
    cursor.execute(sql)  # 執行sql指令
    db.commit()  # 資料同步儲存
    list1 = cursor.fetchall()  # 將資料轉換成陣列
    b = SQLDataPrint(list1)
    write_excel(caseData,b)


    tree.delete(*tree.get_children())                               # 每次讀檔都會將treeview清零
    book = openpyxl.load_workbook('write.xlsx', data_only=True)     # openpyxl讀檔
    sheet = book.active                                             # 讀取工作欄
    contacts = []                                                   # 空矩陣contacts，用來裝給treeview顯示的值
    for row in sheet.rows:                                          # for矩陣，欄位中的row
        x = []                                                      # 空陣列x
        for column in range(sheet.max_column):                      # 因為每筆資料寬度都相同，所以直接用for迴圈跑最大column
            x.append(row[column].value)                             # 將相同row的column資料一筆一筆放入x中
        contacts.append(x)                                          # 將塞好一整筆row的x塞入contacts矩陣中，所以contacts的每一筆資料都是一個row資料
    contacts.pop(0)                                                 # 將header值去掉
    # add data to the treeview
    for contact in contacts:                                        # for矩陣跑contacts
        tree.insert('', tk.END, values=contact)                     # 跑出每一筆contacts資料在treeview上

def insert():                                               #一鍵可以儲存剛剛輸入的值
    list1=[]                                                #建立一個空陣列
    list1.append(spinboxValue1.get())                       #將spinboxValue1放入list1
    list1.append(entry1.get())                              #將entry1放入list1
    list1.append(entry2.get())                              #將entry2放入list1
    list1.append(entry3.get())                              # 將entry2放入list1
    list1.append(entry4.get())                              # 將entry2放入list1
    list1.append(entry5.get())                              # 將entry2放入list1
    print(list1)
    sql1=""
    sql = "INSERT INTO `persons`(`Id`, `LastName`, `FirstName`, `Address`, `City`, `Age`) VALUES "
    for x in range(len(list1)):                                         #for矩陣跑輸入的list1
        sql1=sql1+"'"+list1[x]+"'"              #將輸入的值放入excel欄位中
        if x<len(list1)-1:
            sql1=sql1+','
    sql=sql+"("+sql1+")"
    cursor.execute(sql)  # 執行sql指令
    db.commit()  # 資料同步儲存

    view()


def delete():

    for selected_item in tree.selection():                              #從tree上點選
        item = tree.item(selected_item)
        record = item['values']
        print(record)                                                   #print出來

    sql="DELETE FROM persons WHERE Id = "
    sql=sql+str(record[0])
    cursor.execute(sql)  # 執行sql指令
    db.commit()  # 資料同步儲存

    view()



def replace():
    for selected_item in tree.selection():                              #從tree上點選
        item = tree.item(selected_item)
        record = item['values']
        print(record)                                                   #print出來

    list1 = []  # 建立一個空陣列
    list1.append(spinboxValue1.get())  # 將spinboxValue1放入list1
    list1.append(entry1.get())  # 將entry1放入list1
    list1.append(entry2.get())  # 將entry2放入list1
    list1.append(entry3.get())  # 將entry2放入list1
    list1.append(entry4.get())  # 將entry2放入list1
    list1.append(entry5.get())  # 將entry2放入list1
    print(list1)
    sql="UPDATE `persons` SET "
    sql1=""
    for x in range(len(caseData)):
        sql1=sql1 + "`"+str(caseData[x])+"`" + "=" + "'" + str(list1[x]) + "'"
        if x < len(list1) - 1:
            sql1 = sql1 + ','
    sql=sql+sql1+" WHERE "+"`"+str(caseData[0])+"`" +"="+"'"+str(record[0])+"'"
    print(str(sql))
    cursor.execute(sql)  # 執行sql指令
    db.commit()  # 資料同步儲存
    view()





menubar = tk.Menu(win)    #第一層的下拉選單

filemenu = tk.Menu(menubar)                                     #第二層下拉選單
filemenu.add_command(label="Open")
filemenu.add_command(label="insert",command=insert)                 #儲存並印出輸入的值
filemenu.add_command(label="Exit")
menubar.add_cascade(label="File", menu=filemenu)

filemenu = tk.Menu(menubar)                                     #第二層下拉選單
filemenu.add_command(label="Help")
filemenu.add_command(label="About")
menubar.add_cascade(label="Help", menu=filemenu)
win.config(menu=menubar)

####### 9. toolbar
toolbar = tk.Frame(win,bd=1,  relief=RAISED)
toolbar.pack(side=TOP, fill=X)       # 最上方的 toolbar

def toolbarFunQuit():
    exit()





####### label
labels=[]
for i in range(len(caseData)):
    labels.append(Label(win,text=""+caseData[i]))
    labels[i].place(x=10,y=50+(30*i))
######  3. 1個 SpinBox
# 數字調整 Spinbox
def value_changed():
    print(spinboxValue1.get())

spinboxValue1 = tk.StringVar(value=0)                           #滾動式選單
spin_box1 = ttk.Spinbox(win,from_=0,to=9999999999999999,        #可以選擇從0~9999999999999999
    textvariable=spinboxValue1,)                                #基本數值為0
spin_box1.place(x=80,y=50)                                      #位置


####### 2個 Entry 商品名稱, 金額
entry=[]
entry1=tk.Entry(win)                                            # 新增   LastName 輸入框Entry
entry1.place(x=80,y=80+0*30)                                    # 加入   LastName 元件
entry2=tk.Entry(win)                                            # 新增   FirstName 輸入框Entry
entry2.place(x=80,y=80+1*30)                                    # 加入   FirstName 元件
entry3=tk.Entry(win)                                            # 新增   Address 輸入框Entry
entry3.place(x=80,y=80+2*30)                                    # 加入   Address 元件
entry4=tk.Entry(win)                                            # 新增   City 輸入框Entry
entry4.place(x=80,y=80+3*30)                                    # 加入   City 元件
entry5=tk.Entry(win)                                            # 新增   Age 輸入框Entry
entry5.place(x=80,y=80+4*30)                                    # 加入   Age 元件



#######  1個 Tree 列表   ( 訂單編號, 商品名稱, 金額, 備註,訂單狀況 )  110-GUI-tree-表格-列表.py
# define columns
columns = (''+caseData[0], ''+caseData[1], ''+caseData[2], ''+caseData[3], ''+caseData[4],''+caseData[5])
tree = ttk.Treeview(win, columns=columns,height = 10, show='headings' )
tree.place(x=10,y=250)

# define headings
for x in range(len(caseData)):
    tree.column('' + caseData[x], anchor=CENTER, stretch=NO, width=60)
    tree.heading('' + caseData[x], text='' + caseData[x])


# add data to the treeview

def item_selected(event):
    for selected_item in tree.selection():
        item = tree.item(selected_item)
        record = item['values']
        print(record)

tree.bind('<<TreeviewSelect>>', item_selected)


####SAVE按鈕
button1 = tk.Button(win, text="insert",command=insert)      #insert按鈕
button1.pack()                                              #將值儲存
button1.place(x=500,y=250)                                  #按鈕位置

button2 = tk.Button(win, text='delete', command=delete)     #delete按鈕
button2.pack()                                              #將值儲存
button2.place(x=500,y=300)                                  #按鈕位置

button3 = tk.Button(win, text='replace', command=replace)   #replace按鈕
button3.pack()                                              #將值儲存
button3.place(x=500,y=350)                                  #按鈕位置

button3 = tk.Button(win, text='view', command=view)         #view按鈕
button3.pack()                                              #將值儲存
button3.place(x=500,y=200)                                  #按鈕位置
#################




win.mainloop()          # 最後步驟：程式做無限循環